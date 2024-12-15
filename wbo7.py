import os
import openpyxl
from flask import Flask, render_template_string, request, jsonify
import sqlite3
import random
import webbrowser
import threading

app = Flask(__name__)

EXCEL_PATH = 'ganhadores.xlsx'
DB_PATH = 'usuarios.db'


def abrir_navegador():
    webbrowser.open_new('http://127.0.0.1:5000')


def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                cpf TEXT UNIQUE NOT NULL,
                email TEXT NOT NULL,
                numero_aposta INTEGER NOT NULL
            )
        ''')
        conn.commit()


@app.route('/')
def home():
    return render_template_string(HOME_HTML)


@app.route('/entrar')
def entrar():
    return render_template_string(ENTRAR_HTML)


@app.route('/ganhadores')
def ganhadores():
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT datetime("now"), nome, cpf, email, numero_aposta FROM usuarios')
        ganhadores = cursor.fetchall()
    return render_template_string(GANHADORES_HTML, ganhadores=ganhadores)


@app.route('/numeros_sorteados')
def numeros_sorteados():
    if not os.path.exists(EXCEL_PATH):
        return render_template_string(NUMEROS_SORTEADOS_HTML, ganhadores=[])

    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active

    ganhadores = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Pula a primeira linha (cabeçalho)
        ganhadores.append(row)

    return render_template_string(NUMEROS_SORTEADOS_HTML, ganhadores=ganhadores)


@app.route('/cadastrar', methods=['POST'])
def cadastrar():
    data = request.json
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO usuarios (nome, cpf, email, numero_aposta)
                VALUES (?, ?, ?, ?)
            ''', (data['nome'], data['cpf'], data['email'], data['numeroAposta']))
            conn.commit()
        return jsonify({'message': 'Cadastro realizado com sucesso!'})
    except sqlite3.IntegrityError:
        return jsonify({'message': 'CPF já cadastrado!'}), 400
    except Exception as e:
        return jsonify({'message': f'Erro: {str(e)}'}), 500


@app.route('/sortear', methods=['POST'])
def sortear():
    numero_sorteado = random.randint(00, 99)
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT datetime("now"), nome, cpf, email, numero_aposta
            FROM usuarios WHERE numero_aposta = ?
        ''', (numero_sorteado,))
        ganhadores = cursor.fetchall()

    if ganhadores:
        gerar_excel(ganhadores)
        mensagem = 'Parabéns aos ganhadores!'
    else:
        mensagem = 'Não houve ganhadores.'

    return jsonify({'mensagem': mensagem, 'numero_sorteado': numero_sorteado})


def gerar_excel(ganhadores):
    workbook = openpyxl.Workbook() if not os.path.exists(EXCEL_PATH) else openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active

    if sheet.max_row == 1:
        sheet.append(['Data', 'Nome', 'CPF', 'Email', 'Número de Aposta'])

    for g in ganhadores:
        sheet.append(g)

    workbook.save(EXCEL_PATH)


HOME_HTML = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Home - WBO7</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">
</head>
<body>
    <header class="bg-dark text-white p-4">
        <div class="container d-flex justify-content-between">
            <h1>WBO7</h1>
            <nav>
                <ul class="nav">
                    <li class="nav-item"><a href="/" class="nav-link text-white">Home</a></li>
                    <li class="nav-item"><a href="/entrar" class="nav-link text-white">Cadastrar Palpite</a></li>
                    <li class="nav-item"><a href="/ganhadores" class="nav-link text-white">Registros de Apostas</a></li>
                    <li class="nav-item"><a href="/numeros_sorteados" class="nav-link text-white">Números Sorteados</a></li>

                </ul>
            </nav>
        </div>
    </header>

    <main class="container my-5 text-center">
        <h2>Sorteio Online</h2>
        <button class="btn btn-success mt-3" onclick="realizarSorteio()">Sortear</button>
        <h2 class="mt-4">Resultado:</h2>
        <p id="resultado" class="fs-4"></p>
    </main>

    <script>
        async function realizarSorteio() {
            try {
                const response = await fetch('/sortear', { method: 'POST' });
                const data = await response.json();
                alert(data.mensagem);
                document.getElementById('resultado').textContent = `Número sorteado: ${data.numero_sorteado}`;
            } catch (error) {
                console.error('Erro ao realizar o sorteio:', error);
            }
        }
    </script>
</body>
</html>
'''

ENTRAR_HTML = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cadastrar Palpite - WBO7</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">
</head>
<body>
<header class="bg-dark text-white p-4">
        <div class="container d-flex justify-content-between">
            <h1>WBO7</h1>
            <nav>
                <ul class="nav">
                    <li class="nav-item"><a href="/" class="nav-link text-white">Home</a></li>
                    <li class="nav-item"><a href="/entrar" class="nav-link text-white">Cadastrar Palpite</a></li>
                    <li class="nav-item"><a href="/ganhadores" class="nav-link text-white">Registros de Apostas</a></li>
                    <li class="nav-item"><a href="/numeros_sorteados" class="nav-link text-white">Números Sorteados</a></li>

                </ul>
            </nav>
        </div>
    </header>

<div class="container mt-5" style="max-width: 600px; background-color: #343a40; border: 3px solid #444; border-radius: 30px; box-shadow: 0 0 20px rgba(0, 0, 0, 0.8); padding: 30px; color: #fff;">
    <h2 class="text-center" style="color: #e0e0e0;">Cadastre seu Palpite</h2>
    <form id="formCadastro">
        <div class="mb-3">
            <label for="nome" class="form-label" style="color: #ccc;">Nome Completo</label>
            <input type="text" class="form-control" id="nome" required style="background-color: #444; border-outline-light w-100; border-radius: 15px; box-shadow: inset 0 0 10px rgba(0, 0, 0, 0.6); color: #fff; padding: 12px;">
        </div>
        <div class="mb-3">
            <label for="cpf" class="form-label" style="color: #ccc;">CPF</label>
            <input type="text" class="form-control" id="cpf" required style="background-color: #444; border-outline-light w-100; border-radius: 15px; box-shadow: inset 0 0 10px rgba(0, 0, 0, 0.6); color: #fff; padding: 12px;">
        </div>
        <div class="mb-3">
            <label for="email" class="form-label" style="color: #ccc;">E-mail</label>
            <input type="email" class="form-control" id="email" required style="background-color: #444; border-outline-light w-100; border-radius: 15px; box-shadow: inset 0 0 10px rgba(0, 0, 0, 0.6); color: #fff; padding: 12px;">
        </div>
        <div class="mb-3">
            <label for="numeroAposta" class="form-label" style="color: #ccc;">Número de Aposta</label>
            <input type="number" class="form-control" id="numeroAposta" required style="background-color: #444; border-outline-light w-100; border-radius: 15px; box-shadow: inset 0 0 10px rgba(0, 0, 0, 0.6); color: #fff; padding: 12px;">
        </div>
        <button type="submit" class="btn btn-outline-light w-100" style="border-radius: 15px; box-shadow: 0 5px 10px rgba(255, 255, 255, 0.2);" onmouseover="this.style.backgroundColor='#39ff14'; this.style.color='#000';" onmouseout="this.style.backgroundColor='transparent'; this.style.color='white';">Cadastrar</button>
    </form>
</div>

    <script>
        document.getElementById('formCadastro').addEventListener('submit', async (event) => {
            event.preventDefault();
            const nome = document.getElementById('nome').value;
            const cpf = document.getElementById('cpf').value;
            const email = document.getElementById('email').value;
            const numeroAposta = document.getElementById('numeroAposta').value;

            try {
                const response = await fetch('/cadastrar', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ nome, cpf, email, numeroAposta })
                });

                const data = await response.json();
                alert(data.message);
                document.getElementById('formCadastro').reset();
            } catch (error) {
                console.error('Erro ao cadastrar:', error);
            }
        });
    </script>
</body>
</html>
'''

GANHADORES_HTML = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Ganhadores - WBO7</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">
    <style>
        /* Estilizando o contêiner da tabela com rolagem */
        .table-container {
            max-height: 400px;  /* Defina a altura máxima para a rolagem */
            overflow-y: auto;   /* Ativa a rolagem vertical */
            overflow-x: auto;   /* Ativa a rolagem horizontal se necessário */
        }

        /* Estilo opcional para a barra de rolagem */
        .table-container::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }

        .table-container::-webkit-scrollbar-thumb {
            background-color: #6c757d;
            border-radius: 4px;
        }

        .table-container::-webkit-scrollbar-track {
            background-color: #f1f1f1;
        }
    </style>
</head>
<body>
<header class="bg-dark text-white p-4">
    <div class="container d-flex justify-content-between">
        <h1>WBO7</h1>
        <nav>
            <ul class="nav">
                <li class="nav-item"><a href="/" class="nav-link text-white">Home</a></li>
                <li class="nav-item"><a href="/entrar" class="nav-link text-white">Cadastrar Palpite</a></li>
                <li class="nav-item"><a href="/ganhadores" class="nav-link text-white">Registros de Apostas</a></li>
                <li class="nav-item"><a href="/numeros_sorteados" class="nav-link text-white">Números Sorteados</a></li>

                            </ul>
        </nav>
    </div>
</header>

<div class="container mt-5">
    <h2 class="text-center">Lista de Apostas</h2>

    <!-- Contêiner da tabela com rolagem -->
    <div class="table-container" style="border: 4px solid black; border-radius: 30px; box-shadow: 5px 5px 15px rgba(0, 0, 0, 0.6); padding: 15px;">
        <table class="table table-striped">
            <thead>
                <tr>
                    <th>Data</th>
                    <th>Nome</th>
                    <th>CPF</th>
                    <th>Email</th>
                    <th>Número de Aposta</th>
                </tr>
            </thead>
            <tbody>
                {% for g in ganhadores %}
                <tr>
                    <td>{{ g[0] }}</td>
                    <td>{{ g[1] }}</td>
                    <td>{{ g[2] }}</td>
                    <td>{{ g[3] }}</td>
                    <td>{{ g[4] }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
</div>
</body>
</html>
'''


NUMEROS_SORTEADOS_HTML = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Numeros Sorteados - WBO7</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">
    <style>
        /* Estilizando o contêiner da tabela com rolagem */
        .table-container {
            max-height: 400px;  /* Defina a altura máxima para a rolagem */
            overflow-y: auto;   /* Ativa a rolagem vertical */
            overflow-x: auto;   /* Ativa a rolagem horizontal se necessário */
        }

        /* Estilo opcional para a barra de rolagem */
        .table-container::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }

        .table-container::-webkit-scrollbar-thumb {
            background-color: #6c757d;
            border-radius: 4px;
        }

        .table-container::-webkit-scrollbar-track {
            background-color: #f1f1f1;
        }
    </style>
</head>
<body>
<header class="bg-dark text-white p-4">
    <div class="container d-flex justify-content-between">
        <h1>WBO7</h1>
        <nav>
            <ul class="nav">
                <li class="nav-item"><a href="/" class="nav-link text-white">Home</a></li>
                <li class="nav-item"><a href="/entrar" class="nav-link text-white">Cadastrar Palpite</a></li>
                <li class="nav-item"><a href="/ganhadores" class="nav-link text-white">Registros de Apostas</a></li>
                <li class="nav-item"><a href="/numeros_sorteados" class="nav-link text-white">Números Sorteados</a></li>

            </ul>
        </nav>
    </div>
</header>

<div class="container mt-5">
    <div class="row">
        <!-- Tabela de Histórico de Números Sorteados -->
        <div class="col-md-6">
            <h2 class="text-center">Histórico de Números Sorteados</h2>
            <div class="table-container" style="border: 4px solid black; border-radius: 30px; box-shadow: 5px 5px 15px rgba(0, 0, 0, 0.6); padding: 15px;">
                <table class="table table-striped">
                    <thead>
                        <tr>
                            <th>Data</th>
                            <th>Número de Aposta</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for g in ganhadores %}
                        <tr>
                            <td>{{ g[0] }}</td>
                            <td>{{ g[4] }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>

        <!-- Tabela de Histórico de Ganhadores Sorteados -->
        <div class="col-md-6">
            <h2 class="text-center">Histórico de Ganhadores Sorteados</h2>
            <div class="table-container" style="border: 4px solid black; border-radius: 30px; box-shadow: 5px 5px 15px rgba(0, 0, 0, 0.6); padding: 15px;">
                <table class="table table-striped">
                    <thead>
                        <tr>
                            <th>Data</th>
                            <th>Nome</th>
                            <th>CPF</th>
                            <th>Email</th>
                            <th>Número de Aposta</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for g in ganhadores %}
                        <tr>
                            <td>{{ g[0] }}</td>
                            <td>{{ g[1] }}</td>
                            <td>{{ g[2] }}</td>
                            <td>{{ g[3] }}</td>
                            <td>{{ g[4] }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>
    </div>
</div>

</body>
</html>
'''


if __name__ == '__main__':
    init_db()
    threading.Timer(1.5, abrir_navegador).start()
    app.run(debug=True)
